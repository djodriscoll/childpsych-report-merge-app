#!/usr/bin/env python3
"""
Browser Report Merge App
- Runs locally in a browser at http://127.0.0.1:8765
- Reads Word placeholders in the form «Field_name»
- Creates a web form using clear, VA-friendly labels
- Auto-calculates pronouns and age/time fields
- Generates a populated DOCX with merge values converted to ordinary text
- Generates a completed DOCX and can open Gmail or Outlook Web with a pre-addressed draft.
- Webmail cannot receive an automatic attachment from a local browser page; the generated DOCX must be attached manually.

Uses openpyxl if available to create/update the invoice Excel register.
"""

import html
import json
import os
import hashlib
import secrets
import re
import sys
import threading
import time
import shutil
from email.parser import BytesParser
from email import policy
import urllib.parse
import webbrowser
import zipfile
import copy
import xml.etree.ElementTree as ET
from datetime import date, datetime
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path


OOXML_NAMESPACES = {
    "wpc": "http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas",
    "cx": "http://schemas.microsoft.com/office/drawing/2014/chartex",
    "cx1": "http://schemas.microsoft.com/office/drawing/2015/9/8/chartex",
    "cx2": "http://schemas.microsoft.com/office/drawing/2015/10/21/chartex",
    "cx3": "http://schemas.microsoft.com/office/drawing/2016/5/9/chartex",
    "cx4": "http://schemas.microsoft.com/office/drawing/2016/5/10/chartex",
    "cx5": "http://schemas.microsoft.com/office/drawing/2016/5/11/chartex",
    "cx6": "http://schemas.microsoft.com/office/drawing/2016/5/12/chartex",
    "cx7": "http://schemas.microsoft.com/office/drawing/2016/5/13/chartex",
    "cx8": "http://schemas.microsoft.com/office/drawing/2016/5/14/chartex",
    "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "aink": "http://schemas.microsoft.com/office/drawing/2016/ink",
    "am3d": "http://schemas.microsoft.com/office/drawing/2017/model3d",
    "o": "urn:schemas-microsoft-com:office:office",
    "oel": "http://schemas.microsoft.com/office/2019/extlst",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "m": "http://schemas.openxmlformats.org/officeDocument/2006/math",
    "v": "urn:schemas-microsoft-com:vml",
    "wp14": "http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing",
    "wp": "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing",
    "w10": "urn:schemas-microsoft-com:office:word",
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "w14": "http://schemas.microsoft.com/office/word/2010/wordml",
    "w15": "http://schemas.microsoft.com/office/word/2012/wordml",
    "w16cex": "http://schemas.microsoft.com/office/word/2018/wordml/cex",
    "w16cid": "http://schemas.microsoft.com/office/word/2016/wordml/cid",
    "w16": "http://schemas.microsoft.com/office/word/2018/wordml",
    "w16du": "http://schemas.microsoft.com/office/word/2023/wordml/word16du",
    "w16sdtdh": "http://schemas.microsoft.com/office/word/2020/wordml/sdtdatahash",
    "w16sdtfl": "http://schemas.microsoft.com/office/word/2024/wordml/sdtformatlock",
    "w16se": "http://schemas.microsoft.com/office/word/2015/wordml/symex",
    "wpg": "http://schemas.microsoft.com/office/word/2010/wordprocessingGroup",
    "wpi": "http://schemas.microsoft.com/office/word/2010/wordprocessingInk",
    "wne": "http://schemas.microsoft.com/office/word/2006/wordml",
    "wps": "http://schemas.microsoft.com/office/word/2010/wordprocessingShape",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "pic": "http://schemas.openxmlformats.org/drawingml/2006/picture",
}


def register_ooxml_namespaces():
    # ElementTree otherwise rewrites Word namespace prefixes as ns1/ns2.
    # That can leave mc:Ignorable values referring to prefixes such as w14/w15
    # that no longer exist, which Microsoft Word reports as unreadable content.
    for prefix, uri in OOXML_NAMESPACES.items():
        ET.register_namespace(prefix, uri)

register_ooxml_namespaces()

APP_TITLE = "Report Generation"
HOSTED_MODE = bool(os.environ.get("RENDER") or os.environ.get("PORT"))
HOST = "0.0.0.0" if HOSTED_MODE else "127.0.0.1"
PORT = int(os.environ.get("PORT", "8765"))
DISPLAY_LOCATION = "Hosted Render trial; do not enter real patient/client data" if HOSTED_MODE else f"Runs locally at http://{HOST}:{PORT}; not uploaded to the internet"
PLACEHOLDER_RE = re.compile(r"«([^»]+)»")
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE = BASE_DIR / "Report_Temp.docx"
SUPPLEMENTARY_TEMPLATE = BASE_DIR / "Supplementary_Report_Temp.docx"
OUTPUT_DIR = BASE_DIR / "outputs"
SETTINGS_FILE = Path.home() / ".browser_report_merge_app_settings.json"
AUTH_FILE = BASE_DIR / "auth_config.json"
USERS_FILE = BASE_DIR / "users_config.json"
LOGO_FILE = BASE_DIR / "logo-transparent-svg.svg"
INVOICE_DIR = OUTPUT_DIR / "invoices"
INVOICE_REGISTER = OUTPUT_DIR / "Invoice_Register.xlsx"
SESSIONS = {}
DEFAULT_LOGIN_EMAIL = "drdavid@childpsych.ie"
DEFAULT_LOGIN_PASSWORD = "childpsych"
DEFAULT_VA_EMAIL = "admin@childpsych.ie"
DEFAULT_VA_PASSWORD = "childpsych"
DEFAULT_RECIPIENTS = [
    "drdavid@childpsych.ie",
    "info@childpsych.ie",
    "admin@childpsych.ie",
]

FIELD_ORDER = [
    "Ass_date", "in-person/virtually", "Agency", "Forename", "Surname", "DOB", "Claim_no", "Solicitor_name",
    "Date_accident", "Sex", "Pronoun1", "Pronoun2", "Pronoun3", "Age_today", "Age_accident",
    "Time_from_accident_today", "Date_last_assessment", "Guardian_relationship", "Guardian_name", "Guardian_attend1_rel",
    "Guardian_attend2_rel", "Add1", "Add2", "Add_country", "Add_county", "Add_eircode",
    "Report1_author", "Reprot1_date", "Report2_author", "Report2_date", "Report3_author",
    "Report3_date", "Report4_author", "Report4_date", "Report5_author", "Report5_date",
    "Report6_author", "Report6_date", "Translator_name", "Language",
]

FIELD_LABELS = {
    "Ass_date": "Assessment date",
    "Agency": "Agency / instructing party",
    "in-person/virtually": "Appointment type",
    "Forename": "Claimant forename",
    "Surname": "Claimant surname",
    "DOB": "Claimant date of birth",
    "Claim_no": "Claim number",
    "Solicitor_name": "Solicitor name",
    "Date_accident": "Date of incident / accident",
    "Sex": "Claimant sex",
    "Pronoun1": "Subject pronoun - auto-filled",
    "Pronoun2": "Object pronoun - auto-filled",
    "Pronoun3": "Possessive pronoun - auto-filled",
    "Age_today": "Age at assessment - auto-calculated",
    "Age_accident": "Age at accident - auto-calculated",
    "Time_from_accident_today": "Time from accident to assessment - auto-calculated",
    "Date_last_assessment": "Date of last assessment",
    "Guardian_relationship": "Guardian relationship to claimant",
    "Guardian_name": "Guardian full name",
    "Guardian_attend1_rel": "Guardian attending 1 relationship",
    "Guardian_attend2_rel": "Guardian attending 2 relationship",
    "Add1": "Address line 1",
    "Add2": "Address line 2 / town",
    "Add_county": "County",
    "Add_country": "Country",
    "Add_eircode": "Eircode / postcode",
    "Report1_author": "Report 1 author",
    "Reprot1_date": "Report 1 date",
    "Report2_author": "Report 2 author",
    "Report2_date": "Report 2 date",
    "Report3_author": "Report 3 author",
    "Report3_date": "Report 3 date",
    "Report4_author": "Report 4 author",
    "Report4_date": "Report 4 date",
    "Report5_author": "Report 5 author",
    "Report5_date": "Report 5 date",
    "Report6_author": "Report 6 author",
    "Report6_date": "Report 6 date",
    "Translator_name": "Translator name",
    "Language": "Translator language",
}

SECTIONS = [
    ("Claimant and assessment details", ["Ass_date", "in-person/virtually", "Agency", "Forename", "Surname", "DOB", "Sex", "Age_today"]),
    ("Claim and solicitor details", ["Claim_no", "Solicitor_name", "Date_accident", "Age_accident", "Time_from_accident_today"]),
    ("Supplementary report details", ["Date_last_assessment"]),
    ("Guardian / accompanying adult details", ["Guardian_name", "Guardian_relationship", "Guardian_attend1_rel", "Guardian_attend2_rel"]),
    ("Address", ["Add1", "Add2", "Add_country", "Add_county", "Add_eircode"]),
    ("Pronouns - automatically based on claimant sex", ["Pronoun1", "Pronoun2", "Pronoun3"]),
    ("Reports reviewed", ["Report1_author", "Reprot1_date", "Report2_author", "Report2_date", "Report3_author", "Report3_date", "Report4_author", "Report4_date", "Report5_author", "Report5_date", "Report6_author", "Report6_date"]),
    ("Translator", ["Translator_name", "Language"]),
]

DATE_FIELDS = {"Ass_date", "DOB", "Date_accident", "Reprot1_date", "Report2_date", "Report3_date", "Report4_date", "Report5_date", "Report6_date", "Date_last_assessment"}
AUTO_FIELDS = {"Pronoun1", "Pronoun2", "Pronoun3", "Age_today", "Age_accident", "Time_from_accident_today"}

REQUIRED_FIELDS = {
    "Forename", "Surname", "DOB", "Sex", "Ass_date", "Date_accident",
    "Agency", "Claim_no", "Solicitor_name", "Guardian_name",
    "Guardian_relationship", "Guardian_attend1_rel", "in-person/virtually",
}

FIELD_HELPERS = {
    "Guardian_name": "Full name of the adult by whom the minor is suing.",
    "Guardian_relationship": "Relationship of the adult by whom the minor is suing, e.g. mother, father or legal guardian.",
    "Guardian_attend1_rel": "Relationship of the adult who attended the assessment.",
    "Guardian_attend2_rel": "Optional second attending adult, if applicable.",
    "Report1_author": "Only complete reports actually reviewed. Blank reports are omitted automatically.",
    "Reprot1_date": "Date of Report 1, if Report 1 author is completed.",
    "Report2_author": "Optional. Leave blank if no second report was reviewed.",
    "Report2_date": "Optional. Leave blank if no second report was reviewed.",
    "Report3_author": "Optional. Leave blank if no third report was reviewed.",
    "Report3_date": "Optional. Leave blank if no third report was reviewed.",
    "Report4_author": "Optional. Leave blank if no fourth report was reviewed.",
    "Report4_date": "Optional. Leave blank if no fourth report was reviewed.",
    "Report5_author": "Optional. Leave blank if no fifth report was reviewed.",
    "Report5_date": "Optional. Leave blank if no fifth report was reviewed.",
    "Report6_author": "Optional. Leave blank if no sixth report was reviewed.",
    "Report6_date": "Optional. Leave blank if no sixth report was reviewed.",
    "in-person/virtually": "Select whether the assessment was completed in-person or virtually.",
    "Date_last_assessment": "Required for supplementary reports if the template refers to a previous assessment.",
    "Translator_name": "Only shown when Translator included is Yes.",
    "Language": "Only shown when Translator included is Yes.",
}


RELATIONSHIPS = [
    "", "mother", "father", "parent", "stepmother", "stepfather", "grandmother", "grandfather",
    "aunt", "uncle", "older sister", "older brother", "foster mother", "foster father", "legal guardian",
    "guardian ad litem", "social worker", "other"
]

LANGUAGES = [
    "", "Arabic", "Chinese - Cantonese", "Chinese - Mandarin", "Czech", "Dutch", "English", "French",
    "German", "Hindi", "Irish", "Italian", "Latvian", "Lithuanian", "Polish", "Portuguese",
    "Romanian", "Russian", "Slovak", "Spanish", "Ukrainian", "Urdu", "Other"
]

IRELAND_COUNTIES = [
    "", "Co.Antrim", "Co.Armagh", "Co.Carlow", "Co.Cavan", "Co.Clare", "Co.Cork", "Co.Derry",
    "Co.Donegal", "Co.Down", "Co.Dublin", "Co.Fermanagh", "Co.Galway", "Co.Kerry", "Co.Kildare",
    "Co.Kilkenny", "Co.Laois", "Co.Leitrim", "Co.Limerick", "Co.Longford", "Co.Louth", "Co.Mayo",
    "Co.Meath", "Co.Monaghan", "Co.Offaly", "Co.Roscommon", "Co.Sligo", "Co.Tipperary", "Co.Tyrone",
    "Co.Waterford", "Co.Westmeath", "Co.Wexford", "Co.Wicklow"
]

UK_COUNTIES = [
    "", "Bedfordshire", "Berkshire", "Bristol", "Buckinghamshire", "Cambridgeshire", "Cheshire",
    "City of London", "Cornwall", "County Antrim", "County Armagh", "County Down", "County Fermanagh",
    "County Londonderry", "County Tyrone", "Cumbria", "Derbyshire", "Devon", "Dorset", "Durham",
    "East Riding of Yorkshire", "East Sussex", "Essex", "Gloucestershire", "Greater London", "Greater Manchester",
    "Hampshire", "Herefordshire", "Hertfordshire", "Isle of Wight", "Kent", "Lancashire", "Leicestershire",
    "Lincolnshire", "Merseyside", "Norfolk", "North Yorkshire", "Northamptonshire", "Northumberland",
    "Nottinghamshire", "Oxfordshire", "Rutland", "Shropshire", "Somerset", "South Yorkshire", "Staffordshire",
    "Suffolk", "Surrey", "Tyne and Wear", "Warwickshire", "West Midlands", "West Sussex", "West Yorkshire",
    "Wiltshire", "Worcestershire", "Aberdeenshire", "Angus", "Argyll and Bute", "Ayrshire", "Clackmannanshire",
    "Dumfries and Galloway", "Dunbartonshire", "Edinburgh", "Fife", "Glasgow", "Highland", "Lanarkshire",
    "Midlothian", "Moray", "Orkney", "Perth and Kinross", "Renfrewshire", "Scottish Borders", "Shetland",
    "Stirling", "West Lothian", "Anglesey", "Blaenau Gwent", "Bridgend", "Caerphilly", "Cardiff", "Carmarthenshire",
    "Ceredigion", "Conwy", "Denbighshire", "Flintshire", "Gwynedd", "Merthyr Tydfil", "Monmouthshire",
    "Neath Port Talbot", "Newport", "Pembrokeshire", "Powys", "Rhondda Cynon Taf", "Swansea", "Torfaen",
    "Vale of Glamorgan", "Wrexham"
]



def password_hash(password, salt=None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 200000).hex()
    return {"salt": salt, "hash": digest}


def verify_password(password, stored):
    if not stored or "salt" not in stored or "hash" not in stored:
        return False
    calc = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), stored["salt"].encode("utf-8"), 200000).hex()
    return secrets.compare_digest(calc, stored["hash"])



def default_users_data():
    admin_user = {
        "id": 1,
        "email": DEFAULT_LOGIN_EMAIL,
        "password": password_hash(DEFAULT_LOGIN_PASSWORD),
        "is_admin": True,
        "can_report": True,
        "can_invoice": True,
        "active": True,
    }
    va_user = {
        "id": 2,
        "email": DEFAULT_VA_EMAIL,
        "password": password_hash(DEFAULT_VA_PASSWORD),
        "is_admin": False,
        "can_report": True,
        "can_invoice": True,
        "active": True,
    }
    return {"next_id": 3, "users": [admin_user, va_user]}


def normalise_users_data(data):
    if not isinstance(data, dict):
        data = default_users_data()
    users = data.get("users")
    if not isinstance(users, list):
        data = default_users_data()
        users = data["users"]

    changed = False
    # Ensure each user has the current expected privilege keys.
    for u in users:
        if "active" not in u:
            u["active"] = True; changed = True
        if "is_admin" not in u:
            u["is_admin"] = False; changed = True
        if "can_report" not in u:
            u["can_report"] = True; changed = True
        if "can_invoice" not in u:
            u["can_invoice"] = False; changed = True

    existing = {str(u.get("email", "")).lower(): u for u in users}

    # Ensure the original administrator account exists unless it has been deliberately changed.
    if DEFAULT_LOGIN_EMAIL.lower() not in existing:
        next_id = int(data.get("next_id", 1) or 1)
        admin = default_users_data()["users"][0]
        admin["id"] = next_id
        users.append(admin)
        data["next_id"] = next_id + 1
        changed = True

    # Add the default VA account requested by Dr O'Driscoll.
    if DEFAULT_VA_EMAIL.lower() not in existing:
        next_id = int(data.get("next_id", 1) or 1)
        va = default_users_data()["users"][1]
        va["id"] = next_id
        users.append(va)
        data["next_id"] = next_id + 1
        changed = True

    max_id = max([int(u.get("id", 0) or 0) for u in users] + [0])
    if int(data.get("next_id", 1) or 1) <= max_id:
        data["next_id"] = max_id + 1
        changed = True

    if changed:
        try:
            save_users(data)
        except Exception:
            pass
    return data


def load_users():
    try:
        if USERS_FILE.exists():
            data = json.loads(USERS_FILE.read_text(encoding="utf-8"))
            return normalise_users_data(data)
    except Exception:
        pass

    data = default_users_data()
    try:
        if AUTH_FILE.exists():
            old = json.loads(AUTH_FILE.read_text(encoding="utf-8"))
            if "email" in old and "password" in old:
                data["users"][0]["email"] = old.get("email", DEFAULT_LOGIN_EMAIL)
                data["users"][0]["password"] = old.get("password", password_hash(DEFAULT_LOGIN_PASSWORD))
    except Exception:
        pass
    try:
        save_users(data)
    except Exception:
        pass
    return data


def save_users(data):
    USERS_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")


def find_user_by_email(email):
    email = (email or "").strip().lower()
    for user in load_users().get("users", []):
        if user.get("email", "").lower() == email:
            return user
    return None


def find_user_by_id(user_id):
    try:
        user_id = int(user_id)
    except Exception:
        return None
    for user in load_users().get("users", []):
        if int(user.get("id", 0)) == user_id:
            return user
    return None


def public_user(user):
    return {
        "id": user.get("id"),
        "email": user.get("email", ""),
        "is_admin": bool(user.get("is_admin")),
        "can_report": bool(user.get("can_report", True)),
        "can_invoice": bool(user.get("can_invoice", False)),
        "active": bool(user.get("active", True)),
    }


def new_session(user):
    token = secrets.token_urlsafe(32)
    SESSIONS[token] = {**public_user(user), "created": time.time()}
    return token


def valid_session(headers):
    cookie = headers.get("Cookie", "")
    for part in cookie.split(";"):
        part = part.strip()
        if part.startswith("session="):
            token = part.split("=", 1)[1]
            session = SESSIONS.get(token)
            if session:
                user = find_user_by_id(session.get("id"))
                if user and user.get("active", True):
                    session.update(public_user(user))
                    return token, session
                SESSIONS.pop(token, None)
    return None, None


def user_can(session, permission):
    if not session:
        return False
    if session.get("is_admin"):
        return True
    return bool(session.get(permission, False))

def load_settings():
    if HOSTED_MODE:
        return {}
    try:
        if SETTINGS_FILE.exists():
            return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}


def save_settings(settings):
    if HOSTED_MODE:
        return
    clean = {k: v for k, v in settings.items() if k != "smtp_password"}
    SETTINGS_FILE.write_text(json.dumps(clean, indent=2), encoding="utf-8")


def docx_xml_members(zf):
    for name in zf.namelist():
        if name.startswith("word/") and name.endswith(".xml"):
            yield name


def extract_placeholders(docx_path):
    fields = set()
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(docx_path, "r") as zf:
        for name in docx_xml_members(zf):
            try:
                root = ET.fromstring(zf.read(name))
            except Exception:
                continue
            for para in root.findall(".//w:p", ns):
                texts = [node.text or "" for node in para.findall(".//w:t", ns)]
                if not texts:
                    continue
                paragraph_text = "".join(texts)
                for match in PLACEHOLDER_RE.findall(paragraph_text):
                    clean = match.strip()
                    if clean and "<" not in clean and ">" not in clean:
                        fields.add(clean)
    ordered = [f for f in FIELD_ORDER if f in fields]
    ordered += sorted(fields.difference(ordered), key=lambda x: x.lower())
    return ordered


def safe_filename_part(value):
    value = (value or "").strip()
    value = re.sub(r"[^A-Za-z0-9 _.-]+", "", value)
    value = re.sub(r"\s+", "_", value)
    return value[:60] or "Report"


def parse_date(value):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def years_months_between(start, end):
    if not start or not end or end < start:
        return ""
    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1
    years = months // 12
    rem_months = months % 12
    y_label = "year" if years == 1 else "years"
    m_label = "month" if rem_months == 1 else "months"
    return f"{years} {y_label}, {rem_months} {m_label}"


def age_years_between(start, end):
    if not start or not end or end < start:
        return None
    years = end.year - start.year
    if (end.month, end.day) < (start.month, start.day):
        years -= 1
    return years


def child_gender_descriptor(values):
    sex = values.get("Sex", "")
    if sex == "Male":
        base = "boy"
    elif sex == "Female":
        base = "girl"
    else:
        return "boy/girl"
    dob = parse_date(values.get("DOB"))
    assessment = parse_date(values.get("Ass_date")) or date.today()
    years = age_years_between(dob, assessment)
    if years is not None and 13 <= years < 20:
        return f"teenage {base}"
    return base


def computed_values(values):
    values = dict(values)
    sex = values.get("Sex", "")
    if sex == "Male":
        values["Pronoun1"] = "he"
        values["Pronoun2"] = "him"
        values["Pronoun3"] = "his"
    elif sex == "Female":
        values["Pronoun1"] = "she"
        values["Pronoun2"] = "her"
        values["Pronoun3"] = "her"
    else:
        values["Pronoun1"] = ""
        values["Pronoun2"] = ""
        values["Pronoun3"] = ""

    dob = parse_date(values.get("DOB"))
    accident = parse_date(values.get("Date_accident"))
    assessment = parse_date(values.get("Ass_date"))
    values["Age_today"] = years_months_between(dob, date.today())
    values["Age_accident"] = years_months_between(dob, accident)
    # The Word template labels this as time from accident to examination date, so it is calculated as assessment date minus accident date.
    values["Time_from_accident_today"] = years_months_between(accident, assessment)
    values["Child_gender_descriptor"] = child_gender_descriptor(values)
    return values


def ordinal(n):
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def format_date_ddmmyyyy(value):
    d = parse_date(value)
    return d.strftime("%d/%m/%Y") if d else str(value or "")


def format_date_long(value):
    d = parse_date(value)
    return f"{ordinal(d.day)} {d.strftime('%B')} {d.year}" if d else str(value or "")


def should_use_long_date(field_name, paragraph_text):
    text = (paragraph_text or "").lower()
    if field_name == "Ass_date" and ("guardian on" in text or "based on my examination on" in text):
        return True
    if field_name == "Date_accident" and ("on the" in text or "due to an accident on" in text):
        return True
    if field_name == "Date_last_assessment" and "previous assessment on" in text:
        return True
    return False


def replacement_for_field(field_name, values, paragraph_text=""):
    value = values.get(field_name, "")
    if field_name in DATE_FIELDS:
        return format_date_long(value) if should_use_long_date(field_name, paragraph_text) else format_date_ddmmyyyy(value)
    return str(value or "")


def remove_bold_from_rpr(rpr, ns):
    if rpr is None:
        return None
    for tag in ("b", "bCs"):
        for el in list(rpr.findall(f"w:{tag}", ns)):
            rpr.remove(el)
    return rpr


def front_page_value_paragraph(paragraph_text):
    labels = (
        "Date of assessment:", "Agency:", "Claimant:", "Date of Birth:",
        "Claim number:", "Solicitor:", "Date of incident:"
    )
    return any(label in (paragraph_text or "") for label in labels)


def make_text_run(text, ns, rpr=None, force_not_bold=False):
    w_ns = ns["w"]
    run = ET.Element(f"{{{w_ns}}}r")
    if rpr is not None:
        rpr_copy = copy.deepcopy(rpr)
        if force_not_bold:
            remove_bold_from_rpr(rpr_copy, ns)
        run.append(rpr_copy)
    text_el = ET.SubElement(run, f"{{{w_ns}}}t")
    text_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    text_el.text = text
    return run


def run_text(run, ns):
    return "".join((node.text or "") for node in run.findall(".//w:t", ns) + run.findall(".//w:instrText", ns))


def fld_char_type(run, ns):
    fld = run.find("w:fldChar", ns)
    if fld is None:
        return None
    return fld.attrib.get(f"{{{ns['w']}}}fldCharType")


def visible_placeholder_field_name(field_runs, ns):
    # Some edited templates can have a MERGEFIELD instruction naming one field
    # while the visible result text contains a different «Placeholder». Prefer
    # the visible placeholder because that is what the template author intends.
    after_sep = False
    visible = []
    for r in field_runs:
        ft = fld_char_type(r, ns)
        if ft == "separate":
            after_sep = True
            continue
        if ft == "end":
            break
        if after_sep:
            visible.append(run_text(r, ns))
    match = PLACEHOLDER_RE.search("".join(visible))
    return match.group(1).strip() if match else ""


def replace_merge_fields_in_paragraph(para, values, ns):
    children = list(para)
    para_text = "".join(run_text(child, ns) for child in children)
    changed = False
    i = 0
    while i < len(children):
        child = children[i]
        if child.tag != f"{{{ns['w']}}}r" or fld_char_type(child, ns) != "begin":
            i += 1
            continue
        depth = 0
        end_idx = None
        for j in range(i, len(children)):
            if children[j].tag == f"{{{ns['w']}}}r":
                ft = fld_char_type(children[j], ns)
                if ft == "begin":
                    depth += 1
                elif ft == "end":
                    depth -= 1
                    if depth == 0:
                        end_idx = j
                        break
        if end_idx is None:
            i += 1
            continue
        seq = children[i:end_idx + 1]
        instr = " ".join((n.text or "") for r in seq for n in r.findall(".//w:instrText", ns))
        match = re.search(r"MERGEFIELD\s+([^\\\s]+)", instr)
        if not match:
            i = end_idx + 1
            continue
        field_name = visible_placeholder_field_name(seq, ns) or match.group(1).strip().strip('"')
        replacement = replacement_for_field(field_name, values, para_text)
        rpr = None
        # Prefer the result run properties after the field separator.
        after_sep = False
        for r in seq:
            if fld_char_type(r, ns) == "separate":
                after_sep = True
                continue
            if after_sep and r.find("w:t", ns) is not None:
                rpr = r.find("w:rPr", ns)
                break
        if rpr is None:
            for r in seq:
                if r.find("w:rPr", ns) is not None:
                    rpr = r.find("w:rPr", ns)
                    break
        new_run = make_text_run(replacement, ns, rpr, force_not_bold=front_page_value_paragraph(para_text))
        para.remove(children[i])
        for old in children[i + 1:end_idx + 1]:
            para.remove(old)
        para.insert(i, new_run)
        children = list(para)
        changed = True
        i += 1
    return changed


def replace_simple_merge_fields_in_paragraph(para, values, ns):
    children = list(para)
    para_text = "".join(run_text(child, ns) for child in children)
    changed = False
    for i, child in enumerate(list(children)):
        if child.tag != f"{{{ns['w']}}}fldSimple":
            continue
        instr = child.attrib.get(f"{{{ns['w']}}}instr", "")
        match = re.search(r"MERGEFIELD\s+([^\\\s]+)", instr)
        if not match:
            continue
        visible = "".join(run_text(r, ns) for r in child.findall(".//w:r", ns))
        visible_match = PLACEHOLDER_RE.search(visible)
        field_name = visible_match.group(1).strip() if visible_match else match.group(1).strip().strip('"')
        replacement = replacement_for_field(field_name, values, para_text)
        rpr = None
        for r in child.findall(".//w:r", ns):
            if r.find("w:rPr", ns) is not None:
                rpr = r.find("w:rPr", ns)
                break
        new_run = make_text_run(replacement, ns, rpr, force_not_bold=front_page_value_paragraph(para_text))
        para.remove(child)
        para.insert(i, new_run)
        changed = True
    return changed


def replace_plain_placeholders(root, values, ns):
    changed = False
    # This handles any non-field «Placeholder» text that remains after MERGEFIELD conversion.
    for para in root.findall(".//w:p", ns):
        nodes = para.findall(".//w:t", ns)
        if not nodes:
            continue
        combined = "".join(node.text or "" for node in nodes)
        if "«" not in combined:
            continue
        paragraph_context = combined
        new_text = combined
        for field_name in values:
            new_text = new_text.replace(f"«{field_name}»", replacement_for_field(field_name, values, paragraph_context))
        if new_text != combined:
            nodes[0].text = new_text
            nodes[0].set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            for node in nodes[1:]:
                node.text = ""
            changed = True
    return changed


def replace_boy_girl_text(root, values, ns):
    descriptor = values.get("Child_gender_descriptor") or "boy/girl"
    if descriptor == "boy/girl":
        return False
    changed = False
    for node in root.findall(".//w:t", ns):
        if node.text and "boy/girl" in node.text:
            node.text = node.text.replace("boy/girl", descriptor)
            node.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            changed = True
    return changed


def build_reports_reviewed_text(values):
    pairs = [
        ("Report1_author", "Reprot1_date"),
        ("Report2_author", "Report2_date"),
        ("Report3_author", "Report3_date"),
        ("Report4_author", "Report4_date"),
        ("Report5_author", "Report5_date"),
        ("Report6_author", "Report6_date"),
    ]
    entries = []
    for author_field, date_field in pairs:
        author = (values.get(author_field) or "").strip()
        if not author:
            continue
        d = (values.get(date_field) or "").strip()
        d_text = format_date_ddmmyyyy(d) if d else ""
        entries.append(f"{author} ({d_text})" if d_text else author)
    return f"(Report by {', '.join(entries)})" if entries else ""


def set_paragraph_plain_text(para, text, ns):
    nodes = para.findall(".//w:t", ns)
    if not nodes:
        return False
    nodes[0].text = text
    nodes[0].set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    for node in nodes[1:]:
        node.text = ""
    return True


def tidy_reports_reviewed_paragraphs(root, values, ns):
    reviewed = build_reports_reviewed_text(values)
    changed = False
    for para in root.findall(".//w:p", ns):
        text = paragraph_text(para, ns)
        if "Report by" not in text or "and by interview" not in text:
            continue
        if "available information" not in text:
            continue
        if reviewed:
            new_text = re.sub(r"\(Report by.*?\)\s+and\s+by\s+interview", reviewed + " and by interview", text, flags=re.S)
        else:
            new_text = re.sub(r"\s*\(Report by.*?\)\s+and\s+by\s+interview", " and by interview", text, flags=re.S)
        # In 2.1.1 use claimant forename only, not full name.
        forename = (values.get("Forename") or "").strip()
        surname = (values.get("Surname") or "").strip()
        if forename and surname:
            new_text = new_text.replace(f"claimant {forename} {surname} and", f"claimant {forename} and")
        new_text = re.sub(r"\s{2,}", " ", new_text).strip()
        if new_text != text:
            set_paragraph_plain_text(para, new_text, ns)
            changed = True
    return changed


def clean_word_settings(root, ns):
    # Remove mail-merge data-source settings and automatic field update prompts.
    # These are not needed because the app writes ordinary text into the generated report
    # and builds a static Contents page. Leaving them in causes Word to ask for recipients
    # or to update fields that may refer to external files.
    w_ns = ns["w"]
    settings_tag = f"{{{w_ns}}}settings"
    if root.tag != settings_tag:
        return False
    changed = False
    for tag in ("mailMerge", "updateFields"):
        for el in list(root.findall(f"w:{tag}", ns)):
            root.remove(el)
            changed = True
    return changed

def paragraph_text(para, ns):
    return "".join((node.text or "") for node in para.findall(".//w:t", ns))


def paragraph_style_id(para, ns):
    ppr = para.find("w:pPr", ns)
    if ppr is None:
        return ""
    style = ppr.find("w:pStyle", ns)
    if style is None:
        return ""
    return style.attrib.get(f"{{{ns['w']}}}val", "")


def make_toc_paragraph(level, number, title, ns):
    w_ns = ns["w"]
    p = ET.Element(f"{{{w_ns}}}p")
    ppr = ET.SubElement(p, f"{{{w_ns}}}pPr")
    style = ET.SubElement(ppr, f"{{{w_ns}}}pStyle")
    style.set(f"{{{w_ns}}}val", "TOC2" if level == 2 else "TOC1")

    def add_text_run(txt):
        r = ET.SubElement(p, f"{{{w_ns}}}r")
        t = ET.SubElement(r, f"{{{w_ns}}}t")
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        t.text = txt

    add_text_run(number)
    r_tab = ET.SubElement(p, f"{{{w_ns}}}r")
    ET.SubElement(r_tab, f"{{{w_ns}}}tab")
    add_text_run(title)
    return p


def make_page_break_paragraph(ns):
    w_ns = ns["w"]
    p = ET.Element(f"{{{w_ns}}}p")
    r = ET.SubElement(p, f"{{{w_ns}}}r")
    br = ET.SubElement(r, f"{{{w_ns}}}br")
    br.set(f"{{{w_ns}}}type", "page")
    return p


def refresh_static_contents(root, ns):
    # Rebuild the visible Contents page from the actual Heading 1/Heading 2 paragraphs
    # after merge fields have been populated. This avoids requiring a manual TOC update.
    body = root.find("w:body", ns)
    if body is None:
        return False
    children = list(body)
    paragraphs = [el for el in children if el.tag == f"{{{ns['w']}}}p"]

    contents_para = None
    for p in paragraphs:
        if paragraph_text(p, ns).strip().lower() == "contents":
            contents_para = p
            break
    if contents_para is None:
        return False

    toc_entries = []
    h1 = 0
    h2 = 0
    seen_contents = False
    for p in paragraphs:
        if p is contents_para:
            seen_contents = True
            continue
        if not seen_contents:
            continue
        style = paragraph_style_id(p, ns)
        title = paragraph_text(p, ns).strip()
        if not title:
            continue
        if style == "Heading1":
            h1 += 1
            h2 = 0
            toc_entries.append((1, str(h1), title))
        elif style == "Heading2" and h1:
            h2 += 1
            toc_entries.append((2, f"{h1}.{h2}", title))
    if not toc_entries:
        return False

    # Remove existing TOC field/result paragraphs after the Contents heading.
    current_children = list(body)
    start_idx = current_children.index(contents_para) + 1
    remove = []
    idx = start_idx
    while idx < len(current_children):
        el = current_children[idx]
        if el.tag != f"{{{ns['w']}}}p":
            break
        style = paragraph_style_id(el, ns)
        text = paragraph_text(el, ns).strip()
        if style in {"TOC1", "TOC2", "toc 1", "toc 2"} or not text:
            remove.append(el)
            idx += 1
            continue
        break
    for el in remove:
        body.remove(el)

    insert_at = list(body).index(contents_para) + 1
    for level, number, title in toc_entries:
        body.insert(insert_at, make_toc_paragraph(level, number, title, ns))
        insert_at += 1
    body.insert(insert_at, make_page_break_paragraph(ns))
    return True


def populate_docx(template_path, output_path, values):
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    register_ooxml_namespaces()

    with zipfile.ZipFile(template_path, "r") as zin:
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("word/") and item.filename.endswith(".xml"):
                    try:
                        root = ET.fromstring(data)
                        changed = False
                        for para in root.findall(".//w:p", ns):
                            if replace_merge_fields_in_paragraph(para, values, ns):
                                changed = True
                            if replace_simple_merge_fields_in_paragraph(para, values, ns):
                                changed = True
                        if replace_plain_placeholders(root, values, ns):
                            changed = True
                        if replace_boy_girl_text(root, values, ns):
                            changed = True
                        if item.filename == "word/document.xml" and tidy_reports_reviewed_paragraphs(root, values, ns):
                            changed = True
                        if item.filename == "word/document.xml" and refresh_static_contents(root, ns):
                            changed = True
                        if item.filename == "word/settings.xml":
                            # Always reserialise settings.xml so any older ns1/ns2 namespace
                            # prefixes are normalised back to Word-compatible mc/w14/w15/etc.
                            clean_word_settings(root, ns)
                            changed = True
                        if changed:
                            register_ooxml_namespaces()
                            data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    except Exception:
                        text = data.decode("utf-8", errors="ignore")
                        for field_name in values:
                            text = text.replace(f"«{field_name}»", replacement_for_field(field_name, values, text))
                        data = text.encode("utf-8")
                zout.writestr(item, data)


def webmail_compose_url(provider, recipient, subject, body):
    subject = subject or "Medicolegal report"
    body = body or "Please find attached the completed medicolegal report."
    if provider == "gmail":
        qs = urllib.parse.urlencode({"view": "cm", "fs": "1", "to": recipient, "su": subject, "body": body})
        return "https://mail.google.com/mail/?" + qs
    if provider == "outlook":
        qs = urllib.parse.urlencode({"to": recipient, "subject": subject, "body": body})
        return "https://outlook.office.com/mail/deeplink/compose?" + qs
    qs = urllib.parse.urlencode({"subject": subject, "body": body})
    return f"mailto:{urllib.parse.quote(recipient)}?" + qs


REPORT_CONFIG = {
    "medicolegal": {
        "title": "First Medicolegal Report",
        "subtitle": "",
        "template": DEFAULT_TEMPLATE,
        "filename_prefix": "Medicolegal_Report",
    },
    "supplementary": {
        "title": "Supplementary Medicolegal Report",
        "subtitle": "",
        "template": SUPPLEMENTARY_TEMPLATE,
        "filename_prefix": "Supplementary_Medicolegal_Report",
    },
}
FIELD_CACHE = {}
SETTINGS = load_settings()


def normalise_report_type(report_type):
    return report_type if report_type in REPORT_CONFIG else "medicolegal"


def template_for(report_type):
    report_type = normalise_report_type(report_type)
    template = REPORT_CONFIG[report_type]["template"]
    # While the supplementary template is pending, fall back to a duplicate of the main report template.
    if not template.exists() and report_type == "supplementary":
        return DEFAULT_TEMPLATE
    return template


def fields_for(report_type):
    report_type = normalise_report_type(report_type)
    template = template_for(report_type)
    cache_key = f"{report_type}:{template}:{template.stat().st_mtime if template.exists() else 0}"
    if cache_key not in FIELD_CACHE:
        FIELD_CACHE.clear()
        extracted = extract_placeholders(template) if template.exists() else []
        # The supplementary template uses Date_last_assessment in several headings.
        # Keep it available as a date field even if Word stores one instance as a field result
        # rather than straightforward visible placeholder text.
        if report_type == "supplementary" and "Date_last_assessment" not in extracted:
            insert_after = "Time_from_accident_today"
            if insert_after in extracted:
                extracted.insert(extracted.index(insert_after) + 1, "Date_last_assessment")
            else:
                extracted.append("Date_last_assessment")
        FIELD_CACHE[cache_key] = extracted
    return FIELD_CACHE[cache_key]


def settings_key(report_type, name):
    return f"{normalise_report_type(report_type)}_{name}"


def esc(value):
    return html.escape(str(value or ""), quote=True)


def option_html(options, selected):
    return "".join(f'<option value="{esc(o)}" {"selected" if o == selected else ""}>{esc(o)}</option>' for o in options)


def field_label_html(f, label):
    req = ' <span class="req">*</span>' if f in REQUIRED_FIELDS else ''
    helper = FIELD_HELPERS.get(f, '')
    helper_html = f'<small>{esc(helper)}</small>' if helper else ''
    return f'<span>{esc(label)}{req}</span>{helper_html}'


def render_field(f, saved_values):
    value = saved_values.get(f, "")
    label = FIELD_LABELS.get(f, f.replace("_", " "))
    data_field = esc(f)
    required_attr = ' data-required="1"' if f in REQUIRED_FIELDS else ''
    label_block = field_label_html(f, label)

    if f in AUTO_FIELDS:
        return f"""
        <label class="field readonly-field" data-field-wrap="{data_field}">{label_block}
          <div class="readonly-display" id="display_{data_field}">{esc(value)}</div>
          <input type="hidden" name="field_{data_field}" id="field_{data_field}" value="{esc(value)}">
        </label>"""

    if f == "Sex":
        return f"""
        <label class="field">{label_block}
          <select name="field_{data_field}" id="field_{data_field}"{required_attr}>{option_html(["", "Male", "Female"], value)}</select>
        </label>"""

    if f == "Agency":
        return f"""
        <label class="field">{label_block}
          <select name="field_{data_field}" id="field_{data_field}"{required_attr}>{option_html(["", "Injuries Resolution Board", "Solicitor"], value)}</select>
        </label>"""

    if f == "in-person/virtually":
        return f"""
        <label class="field">{label_block}
          <select name="field_{data_field}" id="field_{data_field}"{required_attr}>{option_html(["", "in-person", "virtually"], value)}</select>
        </label>"""

    if f == "Add_country":
        return f"""
        <label class="field">{label_block}
          <select name="field_{data_field}" id="field_{data_field}"{required_attr}>{option_html(["", "Ireland", "United Kingdom"], value)}</select>
        </label>"""

    if f == "Add_county":
        all_options = sorted(set(IRELAND_COUNTIES + UK_COUNTIES), key=lambda x: x.lower())
        return f"""
        <label class="field">{label_block}
          <select name="field_{data_field}" id="field_{data_field}" data-saved="{esc(value)}"{required_attr}>{option_html(all_options, value)}</select>
        </label>"""

    if f in {"Guardian_relationship", "Guardian_attend1_rel", "Guardian_attend2_rel"}:
        return f"""
        <label class="field">{label_block}
          <select name="field_{data_field}" id="field_{data_field}"{required_attr}>{option_html(RELATIONSHIPS, value)}</select>
        </label>"""

    if f == "Language":
        return f"""
        <label class="field translator-detail">{label_block}
          <select name="field_{data_field}" id="field_{data_field}"{required_attr}>{option_html(LANGUAGES, value)}</select>
        </label>"""

    css_class = "translator-detail" if f == "Translator_name" else ""
    input_type = "date" if f in DATE_FIELDS else "text"
    return f"""
        <label class="field {css_class}">{label_block}
          <input type="{input_type}" name="field_{data_field}" id="field_{data_field}" value="{esc(value)}"{required_attr}>
        </label>"""


SECTION_HINTS = {
    "Claimant and assessment details": "Core identifying information, assessment date and appointment type.",
    "Claim and solicitor details": "Claim number, solicitor and accident timing.",
    "Guardian / accompanying adult details": "Who attended and their relationship to the claimant.",
    "Address": "Postal address details used on the report front sheet.",
    "Pronouns - automatically based on claimant sex": "Locked fields; these are auto-filled from the sex dropdown.",
    "Reports reviewed": "Authors and dates of reports reviewed before assessment.",
    "Translator": "Translator details only appear when translator included is Yes.",
}

def app_shell_styles():
    return """
:root{--navy:#173054;--blue:#225777;--teal:#3a868f;--bg:#f4f8fb;--card:#fff;--ink:#1f2937;--muted:#667085;--line:#d8e1e8;--danger:#b42318;--success:#067647;}
*{box-sizing:border-box} body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;margin:0;background:linear-gradient(180deg,#eef7f9 0,#f7f9fb 320px);color:var(--ink);} 
.header{background:var(--navy);color:#fff;padding:20px 30px;box-shadow:0 4px 18px rgba(16,24,40,.18);} .header-inner{max-width:1120px;margin:0 auto;display:flex;gap:16px;align-items:center;justify-content:space-between}.brand{display:flex;gap:15px;align-items:center}.brand-logo{height:54px;max-width:190px;background:#fff;border-radius:14px;padding:8px}.brand h1{margin:0;font-size:24px}.small{color:#d5ebef;font-size:13px;margin-top:4px}.top-links{display:flex;gap:10px;align-items:center;flex-wrap:wrap}.top-links a{color:#fff;text-decoration:none;font-weight:800;padding:9px 12px;border:1px solid rgba(255,255,255,.25);border-radius:10px}.top-links a:hover{background:rgba(255,255,255,.1)}
main{max-width:1120px;margin:34px auto 70px;padding:0 22px}.card{background:#fff;border:1px solid var(--line);border-radius:22px;box-shadow:0 12px 30px rgba(16,24,40,.07);padding:28px}.hero{display:grid;grid-template-columns:1.2fr .8fr;gap:24px;align-items:center}.hero h2{font-size:30px;margin:0 0 10px;color:var(--navy)}.hero p{color:var(--muted);line-height:1.5;margin:0 0 20px}.big-logo{max-width:320px;width:100%;margin:auto;display:block}.button{display:inline-block;background:var(--blue);color:#fff;border:none;padding:12px 18px;border-radius:12px;text-decoration:none;font-weight:850;box-shadow:0 4px 10px rgba(34,87,119,.22);cursor:pointer}.button.secondary{background:#475467}.button.ghost{background:#fff;color:#344054;border:1px solid #cfd8e3;box-shadow:none}.tile-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:18px}.tile{background:#fff;border:1px solid var(--line);border-radius:20px;box-shadow:0 12px 30px rgba(16,24,40,.07);padding:24px;text-decoration:none;color:var(--ink);display:block;transition:transform .12s ease,box-shadow .12s ease,border-color .12s ease}.tile:hover{transform:translateY(-2px);box-shadow:0 16px 38px rgba(16,24,40,.12);border-color:#9bc8e6}.tile .badge{display:inline-block;background:#e9f5f6;color:var(--teal);padding:5px 10px;border-radius:999px;font-size:12px;font-weight:800;margin-bottom:12px}.tile h3{margin:0 0 18px;font-size:20px}.message{background:#ecfdf3;border:1px solid #abefc6;color:#05603a;padding:13px 15px;border-radius:14px;margin-bottom:16px}.error{background:#fef3f2;border:1px solid #fecdca;color:#b42318;padding:13px 15px;border-radius:14px;margin-bottom:16px}.login-wrap{max-width:440px;margin:54px auto}.login-card h1{margin:0 0 8px;color:var(--navy)}label{display:block;margin:14px 0 6px;font-weight:750;color:#344054}input{width:100%;padding:11px 12px;border:1px solid #cfd8e3;border-radius:10px;font-size:14px}input:focus{outline:none;border-color:var(--blue);box-shadow:0 0 0 4px rgba(34,87,119,.12)}.form-row{margin-bottom:14px}.settings-grid{display:grid;grid-template-columns:1fr 1fr;gap:18px}.note{color:var(--muted);font-size:13px;line-height:1.45}.muted{color:var(--muted)}.home-actions{display:flex;gap:12px;flex-wrap:wrap}.admin-card{margin-top:18px}.inline-form{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:12px;align-items:end}.check{display:flex;gap:8px;align-items:center}.check input{width:auto}.table-wrap{overflow-x:auto}.admin-table{width:100%;border-collapse:collapse;font-size:13px}.admin-table th,.admin-table td{padding:10px;border-bottom:1px solid var(--line);vertical-align:middle}.admin-table th{text-align:left;color:var(--navy)}.admin-table input[type=password]{min-width:180px}.admin-table input[type=checkbox]{width:auto}.success-panel{margin-bottom:18px;display:flex;justify-content:space-between;gap:16px;align-items:center}.actions{display:flex;gap:10px;flex-wrap:wrap}.req{color:var(--danger);font-weight:900}@media(max-width:800px){.hero,.settings-grid{display:block}.brand-logo{height:46px}.header-inner{display:block}.top-links{margin-top:12px}.big-logo{margin-top:22px}.card{padding:22px}}
"""


def logo_img(css_class="brand-logo"):
    return f'<img class="{css_class}" src="/logo.svg" alt="Child Psych logo">'


def login_page(message="", error=""):
    msg_html = f'<div class="message">{esc(message)}</div>' if message else ""
    err_html = f'<div class="error">{esc(error)}</div>' if error else ""
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Login - {APP_TITLE}</title><style>{app_shell_styles()}</style></head>
<body><main class="login-wrap"><section class="card login-card">{logo_img('big-logo')}<h1>Report Generation</h1><p class="muted">Please sign in to continue.</p>{msg_html}{err_html}<form method="post" action="/login"><div class="form-row"><label>Email</label><input type="email" name="email" autocomplete="username" required></div><div class="form-row"><label>Password</label><input type="password" name="password" autocomplete="current-password" required></div><button class="button" type="submit">Sign in</button></form></section></main></body></html>"""



def home_page(message=""):
    msg_html = f'<div class="message">{esc(message)}</div>' if message else ""
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>{APP_TITLE}</title><style>{app_shell_styles()}</style></head>
<body><header class="header"><div class="header-inner"><div class="brand">{logo_img()}<div><h1>Child Psych Home Page</h1><div class="small">{APP_TITLE}</div></div></div><div class="top-links"><a href="/settings">Settings</a><a href="/logout">Sign out</a></div></div></header><main>{msg_html}<section class="card hero"><div><h2>Home</h2><p>Select a task below.</p><div class="home-actions"><a class="button" href="/reports">Report Generation</a><a class="button" href="/invoice">Invoice Entry</a></div></div><div>{logo_img('big-logo')}</div></section></main></body></html>"""

def reports_page(message=""):
    msg_html = f'<div class="message">{esc(message)}</div>' if message else ""
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Reports - {APP_TITLE}</title><style>{app_shell_styles()}</style></head>
<body><header class="header"><div class="header-inner"><div class="brand">{logo_img()}<div><h1>Report Generation</h1><div class="small">Choose report type</div></div></div><div class="top-links"><a href="/">Home</a><a href="/settings">Settings</a><a href="/logout">Sign out</a></div></div></header><main>{msg_html}<section class="tile-grid"><a class="tile" href="/medicolegal"><span class="badge">Main report</span><h3>First Medicolegal Report</h3><span class="button">Generate Report</span></a><a class="tile" href="/supplementary"><span class="badge">Supplementary</span><h3>Supplementary Medicolegal Report</h3><span class="button">Generate Report</span></a></section></main></body></html>"""



def permission_denied_page():
    return home_page("You do not have permission to access that section.")


def settings_page(session=None, message="", error=""):
    session = session or {}
    users_data = load_users()
    current = find_user_by_id(session.get("id")) or {}
    msg_html = f'<div class="message">{esc(message)}</div>' if message else ""
    err_html = f'<div class="error">{esc(error)}</div>' if error else ""
    admin_html = ""
    if session.get("is_admin"):
        rows = []
        for u in users_data.get("users", []):
            uid = int(u.get("id", 0))
            disabled_self = "disabled" if uid == int(session.get("id", 0)) else ""
            checked_admin = "checked" if u.get("is_admin") else ""
            checked_report = "checked" if u.get("can_report") else ""
            checked_invoice = "checked" if u.get("can_invoice") else ""
            checked_active = "checked" if u.get("active", True) else ""
            rows.append(f"""
            <form method="post" action="/settings"><tr>
              <td>{esc(u.get('email',''))}</td>
              <td><input type="checkbox" name="is_admin" value="1" {checked_admin} {disabled_self}></td>
              <td><input type="checkbox" name="can_report" value="1" {checked_report}></td>
              <td><input type="checkbox" name="can_invoice" value="1" {checked_invoice}></td>
              <td><input type="checkbox" name="active" value="1" {checked_active} {disabled_self}></td>
              <td><input type="password" name="new_password" placeholder="Optional reset password"></td>
              <td>
                <input type="hidden" name="user_id" value="{uid}">
                <button class="button" type="submit" name="action" value="admin_update">Save</button>
                <button class="button secondary" type="submit" name="action" value="admin_delete" {disabled_self} onclick="return confirm('Remove this user?');">Remove</button>
              </td>
            </tr></form>""")
        users_table = ''.join(rows)
        admin_html = f"""
        <section class="card admin-card"><h2>Administrator settings</h2><p class="note">Add users by email, set an initial password, and control access to Report Generation and Invoice Entry. Users can change their own password after logging in.</p>
          <div class="admin-add"><h3>Add user</h3><form method="post" action="/settings" class="inline-form"><input type="hidden" name="action" value="admin_add"><label>Email<input type="email" name="email" required></label><label>Initial password<input type="password" name="password" required></label><label class="check"><input type="checkbox" name="can_report" value="1" checked> Report Generation</label><label class="check"><input type="checkbox" name="can_invoice" value="1" checked> Invoice Entry</label><label class="check"><input type="checkbox" name="is_admin" value="1"> Administrator</label><button class="button" type="submit">Add user</button></form></div>
          <h3>Users and privileges</h3><div class="table-wrap"><table class="admin-table"><thead><tr><th>Email</th><th>Admin</th><th>Reports</th><th>Invoices</th><th>Active</th><th>Password reset</th><th>Actions</th></tr></thead><tbody>{users_table}</tbody></table></div>
        </section>"""
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Settings - {APP_TITLE}</title><style>{app_shell_styles()}</style></head>
<body><header class="header"><div class="header-inner"><div class="brand">{logo_img()}<div><h1>Settings</h1><div class="small">Login and administrator settings</div></div></div><div class="top-links"><a href="/">Home</a><a href="/reports">Report Generation</a><a href="/invoice">Invoice Entry</a><a href="/logout">Sign out</a></div></div></header><main>{msg_html}{err_html}<section class="settings-grid"><div class="card"><h2>Your login email</h2><p class="note">Current username: <strong>{esc(current.get('email', session.get('email','')))}</strong></p><form method="post" action="/settings"><input type="hidden" name="action" value="email"><label>New email</label><input type="email" name="new_email" required><label>Confirm new email</label><input type="email" name="confirm_email" required><label>Current password</label><input type="password" name="current_password" required><button class="button" type="submit">Update email</button></form></div><div class="card"><h2>Change your password</h2><form method="post" action="/settings"><input type="hidden" name="action" value="password"><label>Current password</label><input type="password" name="current_password" autocomplete="current-password" required><label>New password</label><input type="password" name="new_password" autocomplete="new-password" required><label>Confirm new password</label><input type="password" name="confirm_password" autocomplete="new-password" required><button class="button" type="submit">Update password</button></form></div></section>{admin_html}</main></body></html>"""

def page(report_type="medicolegal", message="", download_file=""):
    report_type = normalise_report_type(report_type)
    config = REPORT_CONFIG[report_type]
    fields = fields_for(report_type)
    saved_values = SETTINGS.get(settings_key(report_type, "last_values"), {})
    saved_extra = SETTINGS.get(settings_key(report_type, "last_extra"), {})
    translator_included = saved_extra.get("translator_included", "No")

    rendered_fields = {f: render_field(f, saved_values) for f in fields}

    step_sections = []
    shown = set()
    for section_title, section_fields in SECTIONS:
        visible = [rendered_fields[f] for f in section_fields if f in rendered_fields]
        if not visible:
            continue
        shown.update([f for f in section_fields if f in rendered_fields])
        grid_class = "grid"
        body_html = "".join(visible)
        if section_title == "Translator":
            extra = f'''
            <label class="field"><span>Translator included?</span><small>Choose Yes only if a translator was used during the assessment.</small>
              <select name="translator_included" id="translator_included">{option_html(["No", "Yes", "Maybe"], translator_included)}</select>
            </label>'''
            body_html = extra + body_html
        if section_title == "Reports reviewed":
            grid_class = "reports-reviewed-grid"
            rows = []
            report_pairs = [("Report1_author", "Reprot1_date"), ("Report2_author", "Report2_date"), ("Report3_author", "Report3_date"), ("Report4_author", "Report4_date"), ("Report5_author", "Report5_date"), ("Report6_author", "Report6_date")]
            for a, d in report_pairs:
                cells = []
                if a in rendered_fields:
                    cells.append(rendered_fields[a])
                if d in rendered_fields:
                    cells.append(rendered_fields[d])
                if cells:
                    rows.append(f'<div class="report-row">{"".join(cells)}</div>')
            body_html = "".join(rows)
        step_sections.append((section_title, f'<section class="card wizard-step" data-step="{len(step_sections)}"><div class="card-head"><div><h2>{esc(section_title)}</h2><p>{esc(SECTION_HINTS.get(section_title, ""))}</p></div><span class="pill">Step {len(step_sections)+1}</span></div><div class="{grid_class}">{body_html}</div></section>'))

    other_fields = [rendered_fields[f] for f in fields if f not in shown]
    if other_fields:
        title = "Additional details"
        step_sections.append((title, f'<section class="card wizard-step" data-step="{len(step_sections)}"><div class="card-head"><div><h2>{title}</h2><p>Optional fields detected in the selected Word template.</p></div><span class="pill optional">Optional</span></div><div class="grid">{"".join(other_fields)}</div></section>'))

    review_step_index = len(step_sections)
    step_buttons = "".join(f'<button type="button" class="step-nav" data-target="{i}"><span class="step-num">{i+1}</span><span>{esc(title)}</span><span class="step-status" id="stepStatus{i}">○</span></button>' for i, (title, _) in enumerate(step_sections))
    step_buttons += f'<button type="button" class="step-nav" data-target="{review_step_index}"><span class="step-num">{review_step_index+1}</span><span>Review & Generate</span><span class="step-status" id="stepStatus{review_step_index}">○</span></button>'

    sections_html = "".join(html_block for _, html_block in step_sections)

    download = ""
    if download_file:
        download = f'<section class="success-panel"><div><h2>Report generated successfully</h2><p>Filename: <code>{esc(download_file)}</code></p></div><div class="actions"><a class="button" href="/download?file={urllib.parse.quote(download_file)}">Download Word Report</a><a class="button secondary" href="/">Home</a><a class="button secondary" href="/new?type={report_type}">New Report</a></div></section>'

    msg_html = f'<div class="message">{esc(message)}</div>' if message else ""
    ireland_json = json.dumps(IRELAND_COUNTIES)
    uk_json = json.dumps(UK_COUNTIES)
    required_json = json.dumps(sorted([f for f in REQUIRED_FIELDS if f in fields]))
    review_labels_json = json.dumps({f: FIELD_LABELS.get(f, f.replace("_", " ")) for f in fields})

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(config["title"])} - {APP_TITLE}</title>
<style>
:root{{--navy:#173054;--blue:#225777;--teal:#3a868f;--bg:#f3f7fa;--card:#fff;--ink:#1f2937;--muted:#667085;--line:#d8e1e8;--danger:#b42318;--success:#067647;}}
*{{box-sizing:border-box}} body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;margin:0;background:linear-gradient(180deg,#eef7f9 0,#f7f9fb 280px);color:var(--ink);}}
header{{background:var(--navy);color:#fff;padding:16px 28px;position:sticky;top:0;z-index:10;box-shadow:0 4px 18px rgba(16,24,40,.18);}} .header-inner{{max-width:1240px;margin:0 auto;display:flex;justify-content:space-between;align-items:center;gap:16px}}.brand{{display:flex;gap:14px;align-items:center}}.brand-logo{{height:44px;max-width:170px;background:#fff;border-radius:12px;padding:6px}}.top-actions{{display:flex;gap:10px;flex-wrap:wrap}}h1{{margin:0;font-size:22px}}.small{{color:#cde4f4;font-size:13px}}
main{{max-width:1240px;margin:22px auto 70px;padding:0 22px}}.layout{{display:grid;grid-template-columns:295px 1fr;gap:20px;align-items:start}}.sidebar{{position:sticky;top:92px;background:#fff;border:1px solid var(--line);border-radius:18px;box-shadow:0 10px 24px rgba(16,24,40,.06);padding:14px}}.sidebar h2{{font-size:16px;margin:4px 6px 12px}}.step-nav{{width:100%;display:grid;grid-template-columns:28px 1fr 24px;gap:8px;align-items:center;text-align:left;background:#fff;color:var(--ink);box-shadow:none;border:1px solid transparent;padding:10px;border-radius:12px;margin:4px 0;font-weight:700}}.step-nav.active{{background:#eef6fb;border-color:#9bc8e6;color:#12324a}}.step-nav.complete .step-status{{color:var(--success)}}.step-nav.missing .step-status{{color:var(--danger)}}.step-num{{background:#eef4ff;color:#175cd3;border-radius:999px;display:inline-flex;width:24px;height:24px;align-items:center;justify-content:center;font-size:12px}}
.panel,.card{{background:rgba(255,255,255,.96);border:1px solid var(--line);border-radius:18px;box-shadow:0 10px 24px rgba(16,24,40,.06)}}.top-panel{{padding:18px 20px;margin-bottom:16px;display:grid;grid-template-columns:1fr 1fr;gap:18px}}.progress-label{{display:flex;justify-content:space-between;color:var(--muted);font-size:13px;margin-bottom:8px}}.progress-track{{height:12px;background:#e6edf2;border-radius:999px;overflow:hidden}}.progress-bar{{height:100%;width:0%;background:linear-gradient(90deg,var(--blue),#3a9ed4);transition:width .2s ease}}.filename-preview{{font-size:13px;color:var(--muted)}}.filename-preview code{{color:var(--ink);background:#f2f4f7;padding:4px 6px;border-radius:7px}}
.card{{margin:0 0 16px;overflow:hidden}}.wizard-step{{display:none}}.wizard-step.active{{display:block}}.card-head{{display:flex;justify-content:space-between;align-items:center;gap:12px;padding:18px 20px;border-bottom:1px solid #eef2f5}}.card-head h2{{margin:0;font-size:18px}}.card-head p{{margin:4px 0 0;color:var(--muted);font-size:12.5px}}.pill{{white-space:nowrap;background:#eef4ff;color:#175cd3;padding:5px 10px;border-radius:999px;font-size:12px;font-weight:800}}.pill.optional{{background:#f2f4f7;color:#475467}}.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:16px 18px;padding:20px}}.reports-reviewed-grid{{display:flex;flex-direction:column;gap:14px;padding:20px}}.report-row{{display:grid;grid-template-columns:minmax(280px,1fr) 220px;gap:16px;align-items:start;border-bottom:1px solid #eef2f5;padding-bottom:14px}}.report-row:last-child{{border-bottom:none;padding-bottom:0}}
.field span{{display:block;font-weight:700;font-size:13px;margin-bottom:5px;color:#344054}}.field>span:first-child{{min-height:18px}}.req{{color:var(--danger)}} input,select,textarea{{width:100%;padding:10px 11px;border:1px solid #cfd8e3;border-radius:10px;font-size:14px;background:#fff;outline:none;transition:border-color .12s ease,box-shadow .12s ease}} input:focus,select:focus,textarea:focus{{border-color:var(--blue);box-shadow:0 0 0 4px rgba(36,107,159,.12)}}.missing-field input,.missing-field select{{border-color:var(--danger);box-shadow:0 0 0 3px rgba(180,35,24,.08)}}.readonly-display{{width:100%;padding:10px 11px;border:1px solid #d6dde5;border-radius:10px;min-height:41px;background:#f2f5f8;color:#344054}}.readonly-field span:first-child:after{{content:' 🔒';font-weight:400}} small{{display:block;color:var(--muted);font-size:12px;margin-top:5px;font-weight:400}}
.actions{{display:flex;gap:10px;flex-wrap:wrap}}button,.button{{background:var(--blue);color:#fff;border:none;padding:11px 15px;border-radius:11px;text-decoration:none;font-weight:800;cursor:pointer;box-shadow:0 4px 10px rgba(36,107,159,.2)}}button.secondary,.button.secondary{{background:#475467}}button.ghost,.button.ghost{{background:#fff;color:#344054;border:1px solid #cfd8e3;box-shadow:none}}.nav-actions{{display:flex;justify-content:space-between;gap:10px;padding:18px 20px;border-top:1px solid #eef2f5;background:#fbfcfd}}.message{{background:#ecfdf3;border:1px solid #abefc6;color:#05603a;padding:13px 15px;border-radius:14px;margin-bottom:16px}}.warning{{display:none;background:#fffaeb;border:1px solid #fedf89;color:#7a2e0e;padding:13px 15px;border-radius:14px;margin-bottom:16px}}.success-panel{{display:flex;justify-content:space-between;align-items:center;background:#ecfdf3;border:1px solid #abefc6;border-radius:18px;padding:18px 20px;margin:16px 0}}.success-panel h2{{margin:0 0 6px;font-size:18px}}.hidden{{display:none!important}}.review-table{{width:100%;border-collapse:collapse;margin:0}}.review-table th,.review-table td{{border-bottom:1px solid #eef2f5;text-align:left;padding:10px 12px;font-size:14px}}.review-table th{{width:34%;color:#475467;background:#f9fafb}}@media(max-width:900px){{.layout,.top-panel{{display:block}}.report-row{{grid-template-columns:1fr}}.sidebar{{position:static;margin-bottom:16px}}.header-inner{{display:block}}.top-actions{{margin-top:12px}}.success-panel{{display:block}}}}
</style>
</head>
<body>
<header><div class="header-inner"><div class="brand"><img class="brand-logo" src="/logo.svg" alt="Child Psych logo"><div><h1>Report Generation</h1><div class="small">{esc(config["title"])}</div></div></div><div class="top-actions"><a class="button secondary" href="/">Home</a><a class="button secondary" href="/settings">Settings</a><a class="button secondary" href="/new?type={report_type}">New Report</a><button form="reportForm" type="submit" name="action" value="generate">Generate &amp; Download Report</button><a class="button secondary" href="/logout">Sign out</a></div></div></header>
<main>
{msg_html}
{download}
<div id="warningBox" class="warning"></div>
<section class="top-panel panel"><div><h2 style="margin:0 0 8px">{esc(config["title"])}</h2><div class="filename-preview">Output filename: <code id="filenamePreview">Report.docx</code></div></div><div><div class="progress-label"><strong>Required fields completed</strong><span id="completionText">0 of 0</span></div><div class="progress-track"><div class="progress-bar" id="completionBar"></div></div></div></section>
<form id="reportForm" method="post" action="/generate">
  <input type="hidden" name="report_type" value="{esc(report_type)}">
  <div class="layout">
    <aside class="sidebar"><h2>Report steps</h2>{step_buttons}</aside>
    <div>
      {sections_html}
      <section class="card wizard-step" data-step="{review_step_index}"><div class="card-head"><div><h2>Review &amp; Generate</h2><p>Check the key details before downloading the Word report.</p></div><span class="pill">Final step</span></div><div style="padding:20px"><table class="review-table" id="reviewTable"></table></div><div class="nav-actions"><button type="button" class="ghost" id="backBtn2">Back</button><button type="submit" name="action" value="generate">Generate &amp; Download Report</button></div></section>
      <div class="nav-actions" id="wizardActions"><button type="button" class="ghost" id="backBtn">Back</button><div class="actions"><button type="button" class="secondary" id="continueBtn">Continue</button><button type="button" class="ghost" id="reviewBtn">Review</button></div></div>
    </div>
  </div>
</form>
</main>
<script>
const irelandCounties = {ireland_json};
const ukCounties = {uk_json};
const requiredFields = {required_json};
const fieldLabels = {review_labels_json};
const reportPrefix = {json.dumps(config['filename_prefix'])};
const reviewFields = ['Forename','Surname','DOB','Sex','Ass_date','Date_accident','Age_today','Age_accident','Time_from_accident_today','Agency','Claim_no','Solicitor_name','Guardian_name','Guardian_relationship','Guardian_attend1_rel','in-person/virtually','Date_last_assessment'];
let currentStep = 0;
const maxStep = {review_step_index};
function field(id) {{ return document.getElementById('field_' + id); }}
function setAuto(id, value) {{ const input = field(id); const display = document.getElementById('display_' + id); if (input) input.value = value || ''; if (display) display.textContent = value || ''; }}
function parseDate(value) {{ if (!value) return null; const d = new Date(value + 'T00:00:00'); return isNaN(d.getTime()) ? null : d; }}
function yearsMonthsBetween(start, end) {{ if (!start || !end || end < start) return ''; let months = (end.getFullYear() - start.getFullYear()) * 12 + (end.getMonth() - start.getMonth()); if (end.getDate() < start.getDate()) months -= 1; const years = Math.floor(months / 12); const remMonths = months % 12; return `${{years}} ${{years === 1 ? 'year' : 'years'}}, ${{remMonths}} ${{remMonths === 1 ? 'month' : 'months'}}`; }}
function updatePronounsAndAges() {{ const sex = field('Sex') ? field('Sex').value : ''; if (sex === 'Male') {{ setAuto('Pronoun1','he'); setAuto('Pronoun2','him'); setAuto('Pronoun3','his'); }} else if (sex === 'Female') {{ setAuto('Pronoun1','she'); setAuto('Pronoun2','her'); setAuto('Pronoun3','her'); }} else {{ setAuto('Pronoun1',''); setAuto('Pronoun2',''); setAuto('Pronoun3',''); }} const dob=parseDate(field('DOB')?field('DOB').value:''); const accident=parseDate(field('Date_accident')?field('Date_accident').value:''); const assessment=parseDate(field('Ass_date')?field('Ass_date').value:''); setAuto('Age_today', yearsMonthsBetween(dob,new Date())); setAuto('Age_accident', yearsMonthsBetween(dob,accident)); setAuto('Time_from_accident_today', yearsMonthsBetween(accident,assessment)); }}
function updateCountyOptions() {{ const countryEl=field('Add_country'); const countyEl=field('Add_county'); if(!countryEl||!countyEl)return; const existing=countyEl.value||countyEl.dataset.saved||''; let options=['']; if(countryEl.value==='Ireland') options=irelandCounties; else if(countryEl.value==='United Kingdom') options=ukCounties; countyEl.innerHTML=''; options.forEach(opt=>{{const el=document.createElement('option'); el.value=opt; el.textContent=opt; if(opt===existing) el.selected=true; countyEl.appendChild(el);}}); }}
function updateTranslatorVisibility() {{ const sel=document.getElementById('translator_included'); const show=sel&&sel.value==='Yes'; document.querySelectorAll('.translator-detail').forEach(el=>el.classList.toggle('hidden',!show)); if(!show){{ if(field('Translator_name')) field('Translator_name').value=''; if(field('Language')) field('Language').value=''; }} }}
function valueFor(id) {{ const el=field(id); return el ? (el.value||'').trim() : ''; }}
function missingRequired() {{ return requiredFields.filter(id => !valueFor(id)); }}
function updateRequiredHighlights() {{ document.querySelectorAll('.missing-field').forEach(el=>el.classList.remove('missing-field')); missingRequired().forEach(id=>{{ const el=field(id); if(el && el.closest('.field')) el.closest('.field').classList.add('missing-field'); }}); }}
function updateCompletion() {{ const total=requiredFields.length||1; const completed=requiredFields.filter(id=>valueFor(id)).length; const pct=Math.round((completed/total)*100); const t=document.getElementById('completionText'); const b=document.getElementById('completionBar'); if(t)t.textContent=`${{completed}} of ${{total}}`; if(b)b.style.width=pct+'%'; document.querySelectorAll('.step-nav').forEach((btn,i)=>{{ const step=document.querySelector(`.wizard-step[data-step="${{i}}"]`); let requiredInStep=[]; if(step) requiredInStep=requiredFields.filter(id=>{{const el=field(id); return el && step.contains(el);}}); const complete=requiredInStep.length===0 || requiredInStep.every(id=>valueFor(id)); btn.classList.toggle('complete', complete); btn.classList.toggle('missing', !complete); const status=document.getElementById('stepStatus'+i); if(status) status.textContent=complete?'✓':'!'; }}); updateRequiredHighlights(); updateFilenamePreview(); updateReviewTable(); }}
function cleanName(s) {{ return (s||'').replace(/[^A-Za-z0-9 _.-]+/g,'').trim().replace(/\s+/g,'_') || 'Report'; }}
function formatDateForFile(s) {{ return s ? s.replaceAll('-','') : ''; }}
function updateFilenamePreview() {{ const parts=[reportPrefix, cleanName(valueFor('Forename')), cleanName(valueFor('Surname'))]; const claim=cleanName(valueFor('Claim_no')); const ass=formatDateForFile(valueFor('Ass_date')); if(claim && claim!=='Report') parts.push(claim); if(ass) parts.push(ass); const el=document.getElementById('filenamePreview'); if(el) el.textContent=parts.filter(Boolean).join('_')+'.docx'; }}
function updateReviewTable() {{ const table=document.getElementById('reviewTable'); if(!table)return; table.innerHTML=''; reviewFields.forEach(id=>{{ const el=field(id); if(!el) return; const val=valueFor(id); const tr=document.createElement('tr'); const th=document.createElement('th'); const td=document.createElement('td'); th.textContent=fieldLabels[id]||id; td.textContent=val||'—'; tr.appendChild(th); tr.appendChild(td); table.appendChild(tr); }}); }}
function showStep(n) {{ currentStep=Math.max(0,Math.min(maxStep,n)); document.querySelectorAll('.wizard-step').forEach(el=>el.classList.toggle('active', Number(el.dataset.step)===currentStep)); document.querySelectorAll('.step-nav').forEach(el=>el.classList.toggle('active', Number(el.dataset.target)===currentStep)); const actions=document.getElementById('wizardActions'); if(actions) actions.style.display=currentStep===maxStep?'none':'flex'; const back=document.getElementById('backBtn'); if(back) back.disabled=currentStep===0; window.scrollTo({{top:0,behavior:'smooth'}}); updateCompletion(); }}
function showWarning(msg) {{ const box=document.getElementById('warningBox'); if(box){{ box.textContent=msg; box.style.display='block'; }} }}
function clearWarning() {{ const box=document.getElementById('warningBox'); if(box){{ box.textContent=''; box.style.display='none'; }} }}
document.querySelectorAll('.step-nav').forEach(btn=>btn.addEventListener('click',()=>showStep(Number(btn.dataset.target))));
document.getElementById('continueBtn')?.addEventListener('click',()=>showStep(currentStep+1));
document.getElementById('reviewBtn')?.addEventListener('click',()=>showStep(maxStep));
document.getElementById('backBtn')?.addEventListener('click',()=>showStep(currentStep-1));
document.getElementById('backBtn2')?.addEventListener('click',()=>showStep(maxStep-1));
['Sex','DOB','Date_accident','Ass_date'].forEach(id=>{{ if(field(id)) field(id).addEventListener('change', updatePronounsAndAges); }});
if(field('Add_country')) field('Add_country').addEventListener('change', updateCountyOptions);
const translatorSel=document.getElementById('translator_included'); if(translatorSel) translatorSel.addEventListener('change', updateTranslatorVisibility);
document.querySelectorAll('input, select, textarea').forEach(el=>{{ el.addEventListener('input', updateCompletion); el.addEventListener('change', updateCompletion); }});
document.getElementById('reportForm')?.addEventListener('submit', function(e){{ updatePronounsAndAges(); updateCompletion(); const missing=missingRequired(); if(missing.length){{ e.preventDefault(); const labels=missing.map(id=>fieldLabels[id]||id).join(', '); showWarning('Please complete the required fields before generating: '+labels); const first=field(missing[0]); const step=first?first.closest('.wizard-step'):null; if(step) showStep(Number(step.dataset.step)); }} else {{ clearWarning(); }} }});
updateCountyOptions(); updatePronounsAndAges(); updateTranslatorVisibility(); showStep(0); updateCompletion();
</script>
</body>
</html>"""




def parse_multipart_form(body_bytes, content_type):
    """Parse a browser multipart/form-data body without the removed cgi module."""
    raw = (f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n").encode("utf-8") + body_bytes
    msg = BytesParser(policy=policy.default).parsebytes(raw)
    fields = {}
    files = {}
    if not msg.is_multipart():
        return fields, files
    for part in msg.iter_parts():
        disp = part.get("Content-Disposition", "")
        if "form-data" not in disp:
            continue
        name = part.get_param("name", header="content-disposition")
        filename = part.get_filename()
        payload = part.get_payload(decode=True) or b""
        if filename:
            files[name] = {"filename": filename, "content": payload, "content_type": part.get_content_type()}
        else:
            charset = part.get_content_charset() or "utf-8"
            fields[name] = payload.decode(charset, errors="replace")
    return fields, files

INVOICE_FIELDS = [
    ("invoice_date", "Invoice date", "date"),
    ("supplier", "Supplier / payee", "text"),
    ("invoice_number", "Invoice number", "text"),
    ("claimant", "Client / claimant", "text"),
    ("claim_number", "Claim number", "text"),
    ("solicitor_agency", "Solicitor / agency", "text"),
    ("category", "Category", "select"),
    ("amount_ex_vat", "Amount excluding VAT", "number"),
    ("vat", "VAT", "number"),
    ("total", "Total", "number"),
    ("payment_status", "Payment status", "select"),
    ("notes", "Notes", "textarea"),
]
INVOICE_CATEGORIES = ["", "Transcription", "Administration", "Medical report", "Office", "Legal", "Travel", "Software", "Other"]
INVOICE_STATUS = ["", "Unpaid", "Paid", "Queried"]


def invoice_filename(values, original_name="invoice.pdf"):
    d = parse_date(values.get("invoice_date"))
    date_part = d.strftime("%Y-%m-%d") if d else datetime.now().strftime("%Y-%m-%d")
    supplier = safe_filename_part(values.get("supplier", "Supplier"))
    inv = safe_filename_part(values.get("invoice_number", "Invoice"))
    claimant = safe_filename_part(values.get("claimant", "Client"))
    claim = safe_filename_part(values.get("claim_number", "Claim"))
    return f"{date_part}_{supplier}_{inv}_{claimant}_{claim}.pdf"


def update_invoice_register(values, pdf_filename):
    OUTPUT_DIR.mkdir(exist_ok=True)
    headers = ["Date entered", "Invoice date", "Supplier / payee", "Invoice number", "Client / claimant", "Claim number", "Solicitor / agency", "Category", "Amount excluding VAT", "VAT", "Total", "Payment status", "Notes", "PDF filename"]
    row = [datetime.now().strftime("%d/%m/%Y %H:%M"), format_date_ddmmyyyy(values.get("invoice_date")), values.get("supplier", ""), values.get("invoice_number", ""), values.get("claimant", ""), values.get("claim_number", ""), values.get("solicitor_agency", ""), values.get("category", ""), values.get("amount_ex_vat", ""), values.get("vat", ""), values.get("total", ""), values.get("payment_status", ""), values.get("notes", ""), pdf_filename]
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        if INVOICE_REGISTER.exists():
            wb = load_workbook(INVOICE_REGISTER)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice Register"
            ws.append(headers)
            header_fill = PatternFill("solid", fgColor="173054")
            thin = Side(style="thin", color="D8E1E8")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = Border(bottom=thin)
            ws.freeze_panes = "A2"
        ws.append(row)
        widths = [18,14,24,18,24,18,24,18,16,12,12,16,34,38]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        for r in range(2, ws.max_row+1):
            for c in range(1, ws.max_column+1):
                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        for col in (9,10,11):
            for r in range(2, ws.max_row+1):
                ws.cell(r, col).number_format = '€#,##0.00'
        wb.save(INVOICE_REGISTER)
        return True, ""
    except Exception as e:
        import csv
        csv_path = OUTPUT_DIR / "Invoice_Register.csv"
        exists = csv_path.exists()
        with csv_path.open("a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not exists:
                writer.writerow(headers)
            writer.writerow(row)
        return False, str(e)


def render_invoice_field(name, label, ftype):
    if ftype == "select" and name == "category":
        return f'<label class="field"><span>{esc(label)}</span><select name="{name}">{option_html(INVOICE_CATEGORIES, "")}</select></label>'
    if ftype == "select" and name == "payment_status":
        return f'<label class="field"><span>{esc(label)}</span><select name="{name}">{option_html(INVOICE_STATUS, "Unpaid")}</select></label>'
    if ftype == "textarea":
        return f'<label class="field full"><span>{esc(label)}</span><textarea name="{name}" rows="4"></textarea></label>'
    step = ' step="0.01"' if ftype == "number" else ""
    required = " required" if name in {"invoice_date", "supplier", "invoice_number", "total"} else ""
    req = ' <span class="req">*</span>' if required else ""
    return f'<label class="field"><span>{esc(label)}{req}</span><input type="{ftype}" name="{name}"{step}{required}></label>'


def invoice_page(session=None, message="", error="", pdf_file=""):
    msg_html = f'<div class="message">{esc(message)}</div>' if message else ""
    err_html = f'<div class="error">{esc(error)}</div>' if error else ""
    downloads = ""
    if pdf_file:
        downloads = f"""<section class="card success-panel"><div><h2>Invoice prepared</h2><p>Download the renamed PDF and upload it to the correct OneDrive browser folder. Download the Excel register if you want to upload/replace the current register manually.</p><p><code>{esc(pdf_file)}</code></p></div><div class="actions"><a class="button" href="/download_invoice?file={urllib.parse.quote(pdf_file)}">Download renamed PDF</a><a class="button secondary" href="/download_register">Download Excel register</a></div></section>"""
    fields_html = ''.join(render_invoice_field(*f) for f in INVOICE_FIELDS)
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Invoice Entry - {APP_TITLE}</title><style>{app_shell_styles()}</style><style>.invoice-grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px}}.field.full{{grid-column:1/-1}}textarea{{width:100%;padding:11px 12px;border:1px solid #cfd8e3;border-radius:10px;font-size:14px}}.file-box{{background:#f8fbfd;border:1px dashed #9bc8e6;border-radius:16px;padding:18px;margin-bottom:16px}}@media(max-width:760px){{.invoice-grid{{display:block}}}}</style></head>
<body><header class="header"><div class="header-inner"><div class="brand">{logo_img()}<div><h1>Invoice Entry</h1><div class="small">Manual OneDrive workflow for now</div></div></div><div class="top-links"><a href="/">Home</a><a href="/settings">Settings</a><a href="/logout">Sign out</a></div></div></header><main>{msg_html}{err_html}{downloads}<section class="card"><h2>Enter invoice details</h2><p class="note">This prototype renames the uploaded PDF and updates an Excel invoice register for download. OneDrive browser upload remains manual until Microsoft Graph/OneDrive automation is added.</p><form method="post" action="/invoice" enctype="multipart/form-data"><div class="file-box"><label>Invoice PDF <span class="req">*</span><input type="file" name="invoice_pdf" accept="application/pdf,.pdf" required></label></div><div class="invoice-grid">{fields_html}</div><div class="nav-actions"><button class="button" type="submit">Prepare Invoice Files</button></div></form></section></main></body></html>"""

class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        return

    def send_html(self, content, status=200):
        data = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        qs = urllib.parse.parse_qs(parsed.query)
        if parsed.path == "/logo.svg":
            if LOGO_FILE.exists():
                data = LOGO_FILE.read_bytes()
                self.send_response(200)
                self.send_header("Content-Type", "image/svg+xml")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
            self.send_response(404); self.end_headers(); return
        if parsed.path == "/login":
            self.send_html(login_page())
            return
        if parsed.path == "/logout":
            token, _ = valid_session(self.headers)
            if token:
                SESSIONS.pop(token, None)
            self.send_response(302)
            self.send_header("Location", "/login")
            self.send_header("Set-Cookie", "session=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax")
            self.end_headers()
            return
        token, session = valid_session(self.headers)
        if not session:
            self.send_response(302)
            self.send_header("Location", "/login")
            self.end_headers()
            return
        if parsed.path == "/":
            self.send_html(home_page())
            return
        if parsed.path == "/reports":
            if not user_can(session, "can_report"):
                self.send_html(permission_denied_page(), 403)
                return
            self.send_html(reports_page())
            return
        if parsed.path == "/invoice":
            if not user_can(session, "can_invoice"):
                self.send_html(permission_denied_page(), 403)
                return
            self.send_html(invoice_page(session))
            return
        if parsed.path == "/settings":
            self.send_html(settings_page(session))
            return
        if parsed.path in {"/medicolegal", "/supplementary"}:
            if not user_can(session, "can_report"):
                self.send_html(permission_denied_page(), 403)
                return
            report_type = parsed.path.strip("/")
            self.send_html(page(report_type))
            return
        if parsed.path == "/new":
            if not user_can(session, "can_report"):
                self.send_html(permission_denied_page(), 403)
                return
            report_type = normalise_report_type(qs.get("type", ["medicolegal"])[0])
            SETTINGS.pop(settings_key(report_type, "last_values"), None)
            SETTINGS.pop(settings_key(report_type, "last_extra"), None)
            save_settings(SETTINGS)
            self.send_html(page(report_type, "New report started. Previous form values have been cleared."))
            return
        if parsed.path == "/download":
            filename = os.path.basename(qs.get("file", [""])[0])
            path = OUTPUT_DIR / filename
            if not filename or not path.exists():
                self.send_html(page("medicolegal", "File not found."), 404)
                return
            data = path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if parsed.path == "/download_invoice":
            if not user_can(session, "can_invoice"):
                self.send_html(permission_denied_page(), 403)
                return
            filename = os.path.basename(qs.get("file", [""])[0])
            path = INVOICE_DIR / filename
            if not filename or not path.exists():
                self.send_html(invoice_page(session, error="Invoice file not found."), 404)
                return
            data = path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if parsed.path == "/download_register":
            if not user_can(session, "can_invoice"):
                self.send_html(permission_denied_page(), 403)
                return
            if not INVOICE_REGISTER.exists():
                self.send_html(invoice_page(session, error="Invoice register has not been created yet."), 404)
                return
            data = INVOICE_REGISTER.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="Invoice_Register.xlsx"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        self.send_html(home_page())

    def do_POST(self):
        # Invoice PDF uploads use multipart/form-data. Other forms use URL-encoded data.
        if self.path == "/invoice":
            token, session = valid_session(self.headers)
            if not session:
                self.send_response(302); self.send_header("Location", "/login"); self.end_headers(); return
            if not user_can(session, "can_invoice"):
                self.send_html(permission_denied_page(), 403); return
            ctype = self.headers.get("Content-Type", "")
            if "multipart/form-data" not in ctype:
                self.send_html(invoice_page(session, error="Upload a PDF invoice."), 400); return
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            form_fields, form_files = parse_multipart_form(body, ctype)
            values = {name: (form_fields.get(name, "") or "").strip() for name, _, _ in INVOICE_FIELDS}
            pdf_item = form_files.get("invoice_pdf")
            if pdf_item is None or not pdf_item.get("filename"):
                self.send_html(invoice_page(session, error="Please choose a PDF invoice."), 400); return
            original = os.path.basename(pdf_item.get("filename", "invoice.pdf"))
            out_name = invoice_filename(values, original)
            INVOICE_DIR.mkdir(parents=True, exist_ok=True)
            out_path = INVOICE_DIR / out_name
            out_path.write_bytes(pdf_item.get("content", b""))
            ok, err = update_invoice_register(values, out_name)
            msg = "Invoice PDF renamed and invoice register updated." if ok else "Invoice PDF renamed. Excel register could not be created, but a CSV fallback was created."
            self.send_html(invoice_page(session, message=msg, pdf_file=out_name))
            return

        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length).decode("utf-8", errors="replace")
        form = urllib.parse.parse_qs(raw)
        get = lambda name, default="": form.get(name, [default])[0]
        if self.path == "/login":
            email = get("email", "").strip().lower()
            password = get("password", "")
            user = find_user_by_email(email)
            if user and user.get("active", True) and verify_password(password, user.get("password")):
                token = new_session(user)
                self.send_response(302)
                self.send_header("Location", "/")
                self.send_header("Set-Cookie", f"session={token}; Path=/; HttpOnly; SameSite=Lax")
                self.end_headers()
                return
            self.send_html(login_page(error="Incorrect email or password."), 401)
            return
        token, session = valid_session(self.headers)
        if not session:
            self.send_response(302)
            self.send_header("Location", "/login")
            self.end_headers()
            return
        if self.path == "/settings":
            users_data = load_users()
            current_user = find_user_by_id(session.get("id"))
            if not current_user:
                self.send_response(302); self.send_header("Location", "/login"); self.end_headers(); return
            action = get("action")
            if action in {"email", "password"}:
                current = get("current_password")
                if not verify_password(current, current_user.get("password")):
                    self.send_html(settings_page(session, error="Current password is incorrect."), 400)
                    return
                if action == "email":
                    new_email = get("new_email", "").strip().lower()
                    confirm_email = get("confirm_email", "").strip().lower()
                    if not new_email or new_email != confirm_email:
                        self.send_html(settings_page(session, error="The email confirmation does not match."), 400)
                        return
                    existing = find_user_by_email(new_email)
                    if existing and existing.get("id") != current_user.get("id"):
                        self.send_html(settings_page(session, error="That email is already in use."), 400)
                        return
                    for u in users_data["users"]:
                        if u.get("id") == current_user.get("id"):
                            u["email"] = new_email
                    save_users(users_data)
                    session["email"] = new_email
                    self.send_html(settings_page(session, message="Email/username updated."))
                    return
                if action == "password":
                    new_password = get("new_password")
                    confirm_password = get("confirm_password")
                    if len(new_password) < 8:
                        self.send_html(settings_page(session, error="Password must be at least 8 characters."), 400)
                        return
                    if new_password != confirm_password:
                        self.send_html(settings_page(session, error="Password confirmation does not match."), 400)
                        return
                    for u in users_data["users"]:
                        if u.get("id") == current_user.get("id"):
                            u["password"] = password_hash(new_password)
                    save_users(users_data)
                    self.send_html(settings_page(session, message="Password updated."))
                    return
            if not session.get("is_admin"):
                self.send_html(settings_page(session, error="Only administrators can change user privileges."), 403)
                return
            if action == "admin_add":
                email = get("email", "").strip().lower()
                password = get("password", "")
                if not email or len(password) < 8:
                    self.send_html(settings_page(session, error="Enter an email and an initial password of at least 8 characters."), 400)
                    return
                if find_user_by_email(email):
                    self.send_html(settings_page(session, error="That user already exists."), 400)
                    return
                next_id = int(users_data.get("next_id", 1))
                users_data["users"].append({"id": next_id, "email": email, "password": password_hash(password), "is_admin": bool(get("is_admin")), "can_report": bool(get("can_report")), "can_invoice": bool(get("can_invoice")), "active": True})
                users_data["next_id"] = next_id + 1
                save_users(users_data)
                self.send_html(settings_page(session, message="User added."))
                return
            if action == "admin_update":
                uid = int(get("user_id", "0") or 0)
                updated = False
                for u in users_data.get("users", []):
                    if int(u.get("id", 0)) == uid:
                        is_self = uid == int(session.get("id", 0))
                        if not is_self:
                            u["is_admin"] = bool(get("is_admin"))
                            u["active"] = bool(get("active"))
                        u["can_report"] = bool(get("can_report"))
                        u["can_invoice"] = bool(get("can_invoice"))
                        npw = get("new_password", "")
                        if npw:
                            if len(npw) < 8:
                                self.send_html(settings_page(session, error="Password reset must be at least 8 characters."), 400)
                                return
                            u["password"] = password_hash(npw)
                        updated = True
                        break
                if not updated:
                    self.send_html(settings_page(session, error="User not found."), 404)
                    return
                save_users(users_data)
                self.send_html(settings_page(session, message="User privileges updated."))
                return
            if action == "admin_delete":
                uid = int(get("user_id", "0") or 0)
                if uid == int(session.get("id", 0)):
                    self.send_html(settings_page(session, error="You cannot remove your own administrator account while logged in."), 400)
                    return
                users_data["users"] = [u for u in users_data.get("users", []) if int(u.get("id", 0)) != uid]
                save_users(users_data)
                self.send_html(settings_page(session, message="User removed."))
                return
            self.send_html(settings_page(session, error="Unknown settings action."), 400)
            return
        if self.path != "/generate":
            self.send_html(home_page("Unknown action."), 404)
            return
        if not user_can(session, "can_report"):
            self.send_html(permission_denied_page(), 403)
            return
        report_type = normalise_report_type(get("report_type", "medicolegal"))
        fields = fields_for(report_type)
        values = {f: get(f"field_{f}") for f in fields}
        values = computed_values(values)
        translator_included = get("translator_included", "No")
        if translator_included != "Yes":
            values["Translator_name"] = ""
            values["Language"] = ""
        recipient = get("recipient", DEFAULT_RECIPIENTS[0])
        if recipient not in DEFAULT_RECIPIENTS:
            self.send_html(page(report_type, "Invalid recipient selected."), 400)
            return
        OUTPUT_DIR.mkdir(exist_ok=True)
        claimant = safe_filename_part((values.get("Surname", "") + "_" + values.get("Forename", "")).strip("_"))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"{REPORT_CONFIG[report_type]['filename_prefix']}_{claimant}_{timestamp}.docx"
        output_path = OUTPUT_DIR / output_file
        try:
            populate_docx(template_for(report_type), output_path, values)
            if not HOSTED_MODE:
                SETTINGS.update({
                    settings_key(report_type, "recipient"): recipient,
                    settings_key(report_type, "email_subject"): get("email_subject", "Medicolegal report"),
                    settings_key(report_type, "last_values"): values,
                    settings_key(report_type, "last_extra"): {"translator_included": translator_included},
                })
                save_settings(SETTINGS)
            self.send_html(page(report_type, "Report generated successfully.", output_file))
        except Exception as e:
            self.send_html(page(report_type, f"Error: {e}"), 500)


def open_browser_later():
    time.sleep(1)
    webbrowser.open(f"http://{HOST}:{PORT}/")


def main():
    if not DEFAULT_TEMPLATE.exists():
        print(f"Cannot find template: {DEFAULT_TEMPLATE}")
        sys.exit(1)
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"\n{APP_TITLE} is running.")
    if HOSTED_MODE:
        print(f"Render is running the app on port {PORT}.")
    else:
        print(f"Open this address in your browser: http://{HOST}:{PORT}/")
        print("Leave this Terminal window open while using the app.")
        print("Press Ctrl+C in this Terminal window to stop it.\n")
        threading.Thread(target=open_browser_later, daemon=True).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping app.")


if __name__ == "__main__":
    main()
